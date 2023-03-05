import openpyxl
from openpyxl import Workbook
import numpy as np

x1 = openpyxl.load_workbook('Накладная_exel_1.xlsx')
sheet_1 = x1['Лист1']
x2 = openpyxl.load_workbook('Товар складов_exel_2.xlsx')
sheet_2 = x2['Лист1']
x3 = openpyxl.load_workbook('Упаковка_товара_exel_3.xlsx')
sheet_3 = x3['Лист1']
x4 = openpyxl.load_workbook('Расстояние_exel_4.xlsx')
sheet_4 = x4['Лист1']
x5 = openpyxl.load_workbook('Вместимость_exel_5.xlsx')
sheet_5 = x5['Лист1']

class Truck:

    def __init__(self, speed):
        self.speed = speed

    def drive(self):
        start_point = 0
        distance_warehouse_1 = sheet_4['B2'].value
        distance_warehouse_2 = sheet_4['B3'].value
        distance_warehouse_3 = sheet_4['B4'].value
        distance_warehouse_4 = sheet_4['B5'].value
        capacity_truck_sum = 0
        road_to_warehouse = 0
        for i in range(1, 4):
            capacity_truck_sum += sheet_1.cell(row = i + 1, column = 2).value
        if capacity_truck_sum < sheet_5['B2'].value:
            road_to_warehouse = start_point + distance_warehouse_1
        elif capacity_truck_sum < sheet_5['B3'].value:
            road_to_warehouse = start_point + distance_warehouse_2
        elif capacity_truck_sum < sheet_5['B4'].value:
            road_to_warehouse = start_point + distance_warehouse_3
        elif capacity_truck_sum < sheet_5['B5'].value:
            road_to_warehouse = start_point + distance_warehouse_4
        return road_to_warehouse


    def availability_on_the_warehouse_1(self):
        availability = []
        global invoice
        invoice = []
        for v in sheet_1.iter_rows(min_row = 2, max_row = 5, max_col = 1, values_only = True):
            invoice.append(v)
        while sheet_2['A1'].value:
            capacity_b2 = np.array([sheet_5['B2'].value])
            for v in sheet_1.iter_rows(min_row = 2, max_row = 5, max_col = 1, values_only = True):
                for i in sheet_2.iter_rows(min_row = 2, max_row = 4, min_col =1,  max_col = 1, values_only = True):
                    if i == v:
                        for z in sheet_2.iter_rows(min_row = 2, max_row = 4, min_col = 2, max_col = 2, values_only = True):
                            availability.append(z)
                            availability_1 = [sum(i) for i in zip(*availability)]
                            invoice.remove(i)
                            if availability_1 < capacity_b2:
                                return True
                            if invoice != 0:
                                return availability_on_the_warehouse_2()

    def availability_on_the_warehouse_2(self):
        availability = []
        global invoice
        invoice = []
        while sheet_2['C1'].value:
            capacity_b3 = np.array([sheet_5['B3'].value])
            for v in sheet_1.iter_rows(min_row = 2, max_row = 5, max_col = 1, values_only = True):
                for i in sheet_2.iter_rows(min_row = 2, max_row = 4, min_col = 3, max_col = 3, values_only = True):
                    if i == v:
                        for z in sheet_2.iter_rows(min_row = 2, max_row = 4, min_col = 4, max_col = 4,
                                                   values_only = True):
                            availability.append(z)
                            availability_1 = [sum(i) for i in zip(*availability)]
                            if i in invoice:
                                invoice.remove(i)
                            if availability_1 < capacity_b3:
                                return True
                            else:
                                return availability_on_the_warehouse_3()

    def availability_on_the_warehouse_3(self):
        availability = []
        global invoice
        invoice = []
        while sheet_2['E1'].value:
            capacity_b3 = np.array([sheet_5['B4'].value])
            for v in sheet_1.iter_rows(min_row = 2, max_row = 5, max_col = 1, values_only = True):
                for i in sheet_2.iter_rows(min_row = 2, max_row = 5, min_col = 5, max_col = 5, values_only = True):
                    if i == v:
                        for z in sheet_2.iter_rows(min_row = 2, max_row = 5, min_col = 6, max_col = 6,
                                                   values_only = True):
                            availability.append(z)
                            availability_1 = [sum(i) for i in zip(*availability)]
                            if i in invoice:
                                invoice.remove(i)
                            if availability_1 < capacity_b3:
                                return True
                            if invoice != 0:
                                return availability_on_the_warehouse_4()

    def availability_on_the_warehouse_4(self):
        availability = []
        global invoice
        invoice = []
        while sheet_2['G1'].value:
            capacity_b3 = np.array([sheet_5['B5'].value])
            for v in sheet_1.iter_rows(min_row = 2, max_row = 5, max_col = 1, values_only = True):
                for i in sheet_2.iter_rows(min_row = 2, max_row = 4, min_col = 7, max_col = 7, values_only = True):
                    if i == v:
                        for z in sheet_2.iter_rows(min_row = 2, max_row = 5, min_col = 8, max_col = 8,
                                                   values_only = True):
                            availability.append(z)
                            availability_1 = [sum(i) for i in zip(*availability)]
                            if i in invoice:
                                invoice.remove(i)
                            if availability_1 < capacity_b3:
                                return True
                            if invoice != 0:
                                return availability_on_the_warehouse_4()

    def packing_time(self):
        time_packing = []
        if sheet_2['A1'].value:
            speed_packing_wahrehouse = []
            quantity_product = []
            for row_2 in range(2, sheet_3.max_row + 1):
                for k in range(2, sheet_1.max_row + 1):
                    if row_2 == k:
                        speed_packing_wahrehouse.append(sheet_3.cell(row = row_2, column = 2).value)
            for row_1 in range(2, sheet_1.max_row + 1):
                quantity_product.append(sheet_1.cell(row = row_1, column = 2).value)
            time_packing = [speed_packing_wahrehouse[i] * quantity_product[i] for i in
                            range(len(speed_packing_wahrehouse))]
        elif sheet_2['B1'].value:
            speed_packing_wahrehouse = []
            quantity_product = []
            for row_2 in range(2, sheet_3.max_row + 1):
                for k in range(2, sheet_1.max_row + 1):
                    if row_2 == k:
                        speed_packing_wahrehouse.append(sheet_3.cell(row = row_2, column = 4).value)
            for row_1 in range(2, sheet_1.max_row + 1):
                quantity_product.append(sheet_1.cell(row = row_1, column = 2).value)
            time_packing = [speed_packing_wahrehouse[i] * quantity_product[i] for i in
                            range(len(speed_packing_wahrehouse))]
        elif sheet_2['C1'].value:
            speed_packing_wahrehouse = []
            quantity_product = []
            for row_2 in range(2, sheet_3.max_row + 1):
                for k in range(2, sheet_1.max_row + 1):
                    if row_2 == k:
                        speed_packing_wahrehouse.append(sheet_3.cell(row = row_2, column = 6).value)
            for row_1 in range(2, sheet_1.max_row + 1):
                quantity_product.append(sheet_1.cell(row = row_1, column = 2).value)
            time_packing = [speed_packing_wahrehouse[i] * quantity_product[i] for i in
                            range(len(speed_packing_wahrehouse))]
        elif sheet_2['D1'].value:
            speed_packing_wahrehouse = []
            quantity_product = []
            for row_2 in range(2, sheet_3.max_row + 1):
                for k in range(2, sheet_1.max_row + 1):
                    if row_2 == k:
                        speed_packing_wahrehouse.append(sheet_3.cell(row = row_2, column = 8).value)
            for row_1 in range(2, sheet_1.max_row + 1):
                quantity_product.append(sheet_1.cell(row = row_1, column = 2).value)
            time_packing = [speed_packing_wahrehouse[i] * quantity_product[i] for i in
                            range(len(speed_packing_wahrehouse))]
        time_packing_hours = [i / 60 for i in time_packing]
        return round(sum(time_packing_hours), 3)

    def display_packing_time(self):
        if self.packing_time() == 1 or self.packing_time() == 21 or self.packing_time() == 31 or self.packing_time() == 41:
            return 'час.'
        elif self.packing_time() == 0 or 4 < self.packing_time() < 21 or 24 < self.packing_time() < 31 or 34 < self.packing_time() < 41:
            return 'часов.'
        elif (1 < self.packing_time() < 5) or (21 < self.packing_time() < 25) or (31 < self.packing_time() < 35) or (
                41 < self.packing_time() < 45):
            return 'часа.'

    def go_to_startpoint(self):
        full_time_road_to_warehouse = int(self.drive()) * 4
        return full_time_road_to_warehouse

    def visited_warehouses(self):
        if self.availability_on_the_warehouse_1():
            return sheet_2['A1'].value
        elif self.availability_on_the_warehouse_2():
            return sheet_2['B1'].value
        elif self.availability_on_the_warehouse_3():
            return sheet_2['C1'].value
        else:
            return sheet_2['D1'].value

    def save_to_excel_6(self):
        file_excel_6 = Workbook()
        sheet_excel_6 = file_excel_6.create_sheet(title = 'Итоги_excel_6', index = 0)
        sheet_excel_6['A1'] = 'Расстояние выгрузки'
        sheet_excel_6['B1'] = 'Общий путь, пройденный траком'
        sheet_excel_6['C1'] = 'Время упаковки товара'
        sheet_excel_6['D1'] = 'Склад отгрузки'
        sheet_excel_6['E1'] = 'Итоговый отчет'
        sheet_excel_6['A2'] = self.drive()
        sheet_excel_6['B2'] = self.go_to_startpoint()
        sheet_excel_6['C2'] = self.packing_time()
        sheet_excel_6['D2'] = str(self.visited_warehouses())
        sheet_excel_6['E2'] = f"Расстояние до места выгрузки составляет {self.drive()} км, " \
                              f"общее расстояние {self.go_to_startpoint()} км, выгружено на {self.visited_warehouses()}," \
                              f" продолжительность упаковки {self.packing_time()} {self.display_packing_time()}"
        file_excel_6.save(filename = "Итоги_excel_6.xlsx")

    def road_of_truck(self):
        while self.availability_on_the_warehouse_1() == True:
            self.drive()
            self.visited_warehouses()
            self.packing_time()
            self.go_to_startpoint()
            self.save_to_excel_6()
            print(
                f"Расстояние до места выгрузки составляет {self.drive()} км, общее расстояние {self.go_to_startpoint()} км, "
                f"выгружено на {self.visited_warehouses()}, "
                f"продолжительность упаковки {self.packing_time()} {self.display_packing_time()}")
            break
        else:
            print('Склады заполнены, ожидайте разгузки на точке старта.')


if __name__ == "__main__":
    a_truck = Truck(60)
    print(a_truck.road_of_truck())
