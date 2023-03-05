import numpy as np
import openpyxl
from openpyxl import Workbook

x1 = openpyxl.load_workbook('Накладная_exel_1.xlsx')
sheet_1 = x1['Лист1']
x2 = openpyxl.load_workbook('Товар складов_exel_2.xlsx')
sheet_2 = x2['Лист1']
x3 = openpyxl.load_workbook('Упаковка_товара_exel_3.xlsx')
sheet_3 = x3['Лист1']
x4 = openpyxl.load_workbook('Расстояние_exel_4.xlsx')
sheet_4 = x4['Лист1']
x5 = openpyxl.load_workbook('Вместимость_exel_5.xlsx')
sheet_5 = x5['Лист1']


class Truck:
    invoice = []
    def __init__(self, speed):
        self.speed = speed

    def drive(self):
        start_point = 0
        distance_warehouse_1 = sheet_4['B2'].value
        distance_warehouse_2 = sheet_4['B3'].value
        distance_warehouse_3 = sheet_4['B4'].value
        distance_warehouse_4 = sheet_4['B5'].value
        capacity_truck_sum = 0
        road_to_warehouse = 0
        for i in range(1, 4):
            capacity_truck_sum += sheet_1.cell(row = i + 1, column = 2).value
        if capacity_truck_sum < sheet_5['B2'].value:
            road_to_warehouse = start_point + distance_warehouse_1
        elif capacity_truck_sum < sheet_5['B3'].value:
            road_to_warehouse = start_point + distance_warehouse_2
        elif capacity_truck_sum < sheet_5['B4'].value:
            road_to_warehouse = start_point + distance_warehouse_3
        elif capacity_truck_sum < sheet_5['B5'].value:
            road_to_warehouse = start_point + distance_warehouse_4
        return capacity_truck_sum

    def availability_on_the_warehouse_1(self):
        availability = []
        global invoice
        invoice = []
        for v in sheet_1.iter_rows(min_row = 2, max_row = 5, max_col = 1, values_only = True):
            invoice.append(v)
        while sheet_2['A1'].value:
            capacity_b2 = np.array([sheet_5['B2'].value])
            for v in sheet_1.iter_rows(min_row = 2, max_row = 5, max_col = 1, values_only = True):
                for i in sheet_2.iter_rows(min_row = 2, max_row = 4, min_col =1,  max_col = 1, values_only = True):
                    if i == v:
                        for z in sheet_2.iter_rows(min_row = 2, max_row = 4, min_col = 2, max_col = 2, values_only = True):
                            availability.append(z)
                            availability_1 = [sum(i) for i in zip(*availability)]
                            invoice.remove(i)
                            if availability_1 < capacity_b2:
                                return f"Отгружаем на {sheet_2['A1'].value}. {availability_1} {invoice}"
                            else:
                                availability_on_the_warehouse_2()

    def availability_on_the_warehouse_2(self):
        availability = []
        global invoice
        invoice = []
        while sheet_2['C1'].value:
            capacity_b3 = np.array([sheet_5['B3'].value])
            for v in sheet_1.iter_rows(min_row = 2, max_row = 5, max_col = 1, values_only = True):
                for i in sheet_2.iter_rows(min_row = 2, max_row = 4, min_col = 3, max_col = 3, values_only = True):
                    if i == v:
                        for z in sheet_2.iter_rows(min_row = 2, max_row = 4, min_col = 4, max_col = 4,
                                                   values_only = True):
                            availability.append(z)
                            availability_1 = [sum(i) for i in zip(*availability)]
                            if i in invoice:
                                invoice.remove(i)
                            if availability < capacity_b3:
                                return f"Отгружаем на {sheet_2['C1'].value}. {availability_1} {invoice}"
                            else:
                                return availability_on_the_warehouse_3()

    def availability_on_the_warehouse_3(self):
        availability = []
        global invoice
        invoice = []
        while sheet_2['E1'].value:
            capacity_b3 = np.array([sheet_5['B4'].value])
            for v in sheet_1.iter_rows(min_row = 2, max_row = 5, max_col = 1, values_only = True):
                for i in sheet_2.iter_rows(min_row = 2, max_row = 5, min_col = 5, max_col = 5, values_only = True):
                    if i == v:
                        for z in sheet_2.iter_rows(min_row = 2, max_row = 5, min_col = 6, max_col = 6,
                                                   values_only = True):
                            availability.append(z)
                            availability_1 = [sum(i) for i in zip(*availability)]
                            if i in invoice:
                                invoice.remove(i)
                            if availability < capacity_b3:
                                return f"Отгружаем на {sheet_2['E1'].value}. {availability_1} {invoice}"
                            else:
                                return availability_on_the_warehouse_4()

    def availability_on_the_warehouse_4(self):
        availability = []
        global invoice
        invoice = []
        while sheet_2['G1'].value:
            capacity_b3 = np.array([sheet_5['B5'].value])
            for v in sheet_1.iter_rows(min_row = 2, max_row = 5, max_col = 1, values_only = True):
                for i in sheet_2.iter_rows(min_row = 2, max_row = 4, min_col = 7, max_col = 7, values_only = True):
                    if i == v:
                        for z in sheet_2.iter_rows(min_row = 2, max_row = 5, min_col = 8, max_col = 8,
                                                   values_only = True):
                            availability.append(z)
                            availability_1 = [sum(i) for i in zip(*availability)]
                            if i in invoice:
                                invoice.remove(i)
                            if availability < capacity_b3:
                                return f"Отгружаем на {sheet_2['G1'].value}. {availability_1} {invoice}"
                            else:
                                return availability_on_the_warehouse_4()

a_truck = Truck(60)
print(a_truck.availability_on_the_warehouse_4())